import os
import sys
import shutil
import time
import ctypes
from datetime import date, datetime
from pathlib import Path
from typing import List, Tuple, Optional

from PySide6.QtCore import Qt, QDate, QPoint
from PySide6.QtGui import QAction, QIcon, QPainter, QPixmap, QColor, QKeySequence
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QRadioButton, QButtonGroup, QDateEdit, QTextEdit, QTableWidget, QTableWidgetItem,
    QAbstractItemView, QHeaderView, QMessageBox, QMenu, QStatusBar, QFrame, QLineEdit,
    QFileDialog
)

# Excel
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("This tool requires 'openpyxl'. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

APP_TITLE = "Network Changes Tracker"
FILE_NAME = "network_changes.xlsx"
SHEETS = ("CR", "WP")
HEADERS = ("Approval Date", "Request Number", "Description of Work")
DATE_NUMBER_FORMAT = "yyyy-mm-dd"

# UI colours
ACCENT = "#0E9AA7"
BG_LIGHT = "#F6FAFB"
ROW_ALT = "#F2F8F9"
TEXT_PRIMARY = "#0B3C49"


def is_frozen() -> bool:
    return getattr(sys, "frozen", False)


def resource_path(*relative) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, *relative)


def app_dir() -> Path:
    if is_frozen():
        return Path(sys.executable).resolve().parent
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


# -------- Local data directory (keep live file out of OneDrive) --------

def get_data_dir() -> Path:
    if sys.platform == "win32":
        base = os.getenv("LOCALAPPDATA")
        if base:
            return Path(base) / "NetworkChangesTracker"
        # Fallback
        return Path.home() / "AppData" / "Local" / "NetworkChangesTracker"
    elif sys.platform == "darwin":
        return Path.home() / "Library" / "Application Support" / "NetworkChangesTracker"
    else:
        return Path.home() / ".local" / "share" / "NetworkChangesTracker"


DATA_DIR = get_data_dir()
DATA_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = DATA_DIR / FILE_NAME


# -------- OneDrive helper for export --------

def find_onedrive_dir() -> Path:
    # Try environment variables first (Windows)
    for env_name in ("OneDriveCommercial", "OneDriveConsumer", "OneDrive"):
        val = os.getenv(env_name)
        if val and Path(val).exists():
            return Path(val)

    # Common locations
    home = Path.home()
    candidates = [
        home / "OneDrive",
        home / "OneDrive - Personal",
        home / "OneDrive - Microsoft",
        home / "Library" / "CloudStorage",  # macOS OneDrive under CloudStorage
    ]

    # Scan macOS CloudStorage for OneDrive-* folders
    cloud_root = home / "Library" / "CloudStorage"
    if cloud_root.exists():
        for p in cloud_root.iterdir():
            if p.is_dir() and p.name.startswith("OneDrive"):
                candidates.insert(0, p)

    for p in candidates:
        if p.exists():
            return p
    # Fallback to home if nothing found
    return home


# -------- Robust save helpers (atomic save + simple lock + Excel lock detect) --------

def excel_lock_exists(path: Path) -> bool:
    # Excel creates a lock file like "~$network_changes.xlsx" while open
    lock = path.with_name(f"~${path.name}")
    return lock.exists()


def acquire_file_lock(path: Path, timeout: float = 10.0) -> Optional[Tuple[Path, int]]:
    """
    Try to acquire a simple lock using a .lock file next to 'path'.
    Returns (lock_path, fd) if acquired, or None on timeout.
    """
    lock_path = path.with_suffix(path.suffix + ".lock")
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.write(fd, f"pid={os.getpid()} time={time.time()}".encode("utf-8"))
            return lock_path, fd
        except FileExistsError:
            time.sleep(0.2)
        except Exception:
            time.sleep(0.2)
    return None


def release_file_lock(lock: Optional[Tuple[Path, int]]):
    if not lock:
        return
    lock_path, fd = lock
    try:
        try:
            os.close(fd)
        except Exception:
            pass
        if lock_path.exists():
            try:
                lock_path.unlink()
            except Exception:
                pass
    except Exception:
        pass


def safe_save_workbook(wb, path: Path):
    """
    Atomic save: write to a tmp file in the same directory then replace.
    """
    tmp = path.with_suffix(path.suffix + f".tmp.{os.getpid()}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(tmp)
    os.replace(tmp, path)


def save_wb_with_lock(wb, path: Path):
    if excel_lock_exists(path):
        raise PermissionError(
            f"Cannot save the Excel file because it is open in Excel.\n\n"
            f"Please close the workbook first:\n{path}"
        )
    lock = acquire_file_lock(path, timeout=10.0)
    if not lock:
        raise PermissionError(
            f"File appears busy. Try again in a moment.\n\n{path}\n\n"
            "If the file is in OneDrive, ensure it is 'Always keep on this device' and not actively syncing."
        )
    try:
        safe_save_workbook(wb, path)
    finally:
        release_file_lock(lock)


def _migrate_two_to_three_columns(ws) -> bool:
    """
    If the sheet has old layout:
      A1="Approval Date", B1="Description of Work", C1 empty
    migrate by moving B->C for all data rows and inserting B1="Request Number".
    Returns True if migration performed.
    """
    try:
        a1 = (ws["A1"].value or "").strip()
        b1 = (ws["B1"].value or "").strip()
        c1 = ws["C1"].value
        if a1 == "Approval Date" and b1 == "Description of Work" and (c1 is None or str(c1).strip() == ""):
            # Only migrate if column C is empty across data rows to avoid overwriting user data.
            has_c_data = any((ws.cell(row=r, column=3).value not in (None, "")) for r in range(2, ws.max_row + 1))
            if has_c_data:
                # Do not migrate to avoid data loss; just set headers (keeps old data in B).
                ws["B1"] = "Request Number"
                ws["C1"] = "Description of Work"
                return False
            # Move data B -> C
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=3).value = ws.cell(row=r, column=2).value
                ws.cell(row=r, column=2).value = None
            # Set headers
            ws["B1"] = "Request Number"
            ws["C1"] = "Description of Work"
            return True
    except Exception:
        pass
    return False


def ensure_workbook_and_sheets(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    dirty = False
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        dirty = True
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws["A1"] = HEADERS[0]
            ws["B1"] = HEADERS[1]
            ws["C1"] = HEADERS[2]
            dirty = True
        else:
            ws = wb[sheet_name]
            a1 = ws["A1"].value
            b1 = ws["B1"].value
            c1 = ws["C1"].value
            # If all headers empty, write headers
            if (a1 is None) and (b1 is None) and (c1 is None):
                ws["A1"] = HEADERS[0]
                ws["B1"] = HEADERS[1]
                ws["C1"] = HEADERS[2]
                dirty = True
            else:
                # Attempt migration if needed
                migrated = _migrate_two_to_three_columns(ws)
                if migrated:
                    dirty = True
                else:
                    # Ensure header text matches (don’t touch user data positions)
                    changed = False
                    if (ws["A1"].value or "").strip() == "":
                        ws["A1"] = HEADERS[0]
                        changed = True
                    if (ws["B1"].value or "").strip() == "":
                        ws["B1"] = HEADERS[1]
                        changed = True
                    if (ws["C1"].value or "").strip() == "":
                        ws["C1"] = HEADERS[2]
                        changed = True
                    if changed:
                        dirty = True
    if dirty:
        save_wb_with_lock(wb, path)
    return wb


def normalize_description(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = [line.strip() for line in text.split("\n")]
    parts = [p for p in parts if p]
    return ", ".join(parts)


def append_row(path: Path, sheet_name: str, d: date, req_num: str, desc: str):
    wb = ensure_workbook_and_sheets(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws["A1"] = HEADERS[0]
        ws["B1"] = HEADERS[1]
        ws["C1"] = HEADERS[2]
    ws = wb[sheet_name]
    ws.append([d, req_num, desc])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).number_format = DATE_NUMBER_FORMAT
    try:
        save_wb_with_lock(wb, path)
    except PermissionError:
        raise PermissionError(
            f"Cannot save the Excel file.\n\nFile may be open or folder not writable:\n{path}\n\n"
            "Close the file if open, or move the EXE and Excel to a writable folder (e.g., Desktop/Documents)."
        )


def read_rows(path: Path, sheet_name: str) -> List[Tuple[str, str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows: List[Tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        approval_date, req_num, description = (row + (None, None, None))[:3]
        if approval_date is None and req_num is None and description is None:
            continue
        if isinstance(approval_date, (datetime, date)):
            dstr = approval_date.strftime("%Y-%m-%d")
        elif approval_date:
            dstr = str(approval_date)
        else:
            dstr = ""
        rows.append((dstr, str(req_num or ""), description or ""))
    return rows


def emoji_icon(emoji: str, size: int = 128,
               bg=QColor(14, 154, 167), fg=QColor(255, 255, 255)) -> QIcon:
    pm = QPixmap(size, size)
    pm.fill(Qt.transparent)
    painter = QPainter(pm)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setBrush(bg)
    painter.setPen(Qt.NoPen)
    painter.drawEllipse(0, 0, size, size)
    painter.setPen(fg)
    painter.drawText(pm.rect(), Qt.AlignCenter, emoji)
    painter.end()
    return QIcon(pm)


def load_window_icon() -> QIcon:
    # Try common icon names next to the script/exe
    for name in ("app.ico", "app.png", "app.icns"):
        p = app_dir() / name
        if p.exists():
            return QIcon(str(p))
    # Fallback
    return emoji_icon("📡")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(980, 680)
        self.setWindowIcon(load_window_icon())

        ensure_workbook_and_sheets(EXCEL_PATH)

        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # Banner
        banner = QFrame()
        banner.setStyleSheet(f"background:{ACCENT}; border-radius:6px;")
        bl = QVBoxLayout(banner)
        title = QLabel(APP_TITLE)
        subtitle = QLabel("Track network changes quickly and consistently")
        for w in (title, subtitle):
            w.setStyleSheet("color:white;")
        title.setStyleSheet("color:white; font-size:18px; font-weight:600;")
        subtitle.setStyleSheet("color:#E6FFFF; font-size:12px;")
        bl.addWidget(title)
        bl.addWidget(subtitle)
        root.addWidget(banner)

        # Top row: tracker + date + request number
        top = QHBoxLayout()
        lbl_tracker = QLabel("Tracker:")
        self.rb_cr = QRadioButton("CR")
        self.rb_wp = QRadioButton("WP")
        self.rb_cr.setChecked(True)
        self.sheet_group = QButtonGroup(self)
        self.sheet_group.addButton(self.rb_cr)
        self.sheet_group.addButton(self.rb_wp)
        self.sheet_group.buttonToggled.connect(lambda *_: self.on_sheet_changed())

        top.addWidget(lbl_tracker)
        top.addWidget(self.rb_cr)
        top.addWidget(self.rb_wp)
        top.addSpacing(16)

        lbl_date = QLabel("Approval Date:")
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        btn_today = QPushButton("Today")
        btn_today.clicked.connect(self.set_today)

        top.addWidget(lbl_date)
        top.addWidget(self.date_edit)
        top.addWidget(btn_today)
        top.addSpacing(16)

        lbl_req = QLabel("Request Number:")
        self.req_edit = QLineEdit()
        self.req_edit.setPlaceholderText("e.g. CR/ENP/1234 or NC12345678")
        self.req_edit.setMinimumWidth(200)

        top.addWidget(lbl_req)
        top.addWidget(self.req_edit)
        top.addStretch(1)
        root.addLayout(top)

        # Description
        desc_frame = QVBoxLayout()
        lbl_desc = QLabel("Description of Work (multi-line allowed)")
        self.desc_text = QTextEdit()
        self.desc_text.textChanged.connect(self.update_preview)
        self.desc_text.setPlaceholderText("Enter work description; multiple lines will be joined with commas")
        desc_frame.addWidget(lbl_desc)
        desc_frame.addWidget(self.desc_text)
        root.addLayout(desc_frame)

        # Preview
        lbl_prev = QLabel("Preview (single line):")
        self.preview = QLabel("(nothing yet)")
        self.preview.setStyleSheet("background:white; border:1px solid #D0E3E6; padding:8px;")
        root.addWidget(lbl_prev)
        root.addWidget(self.preview)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_add = QPushButton("Add")
        self.btn_add.clicked.connect(self.on_add)
        self.btn_clear = QPushButton("Clear")
        self.btn_clear.clicked.connect(self.on_clear)
        self.btn_open = QPushButton("Open Excel")
        self.btn_open.clicked.connect(self.open_excel)

        # New: Export to OneDrive
        self.btn_export = QPushButton("Export to OneDrive…")
        self.btn_export.setToolTip("Copy the current workbook to your OneDrive folder")
        self.btn_export.clicked.connect(self.on_export_to_onedrive)

        self.btn_refresh = QPushButton("Refresh List")
        self.btn_refresh.clicked.connect(self.load_table)

        tip = QLabel("Shortcuts: Ctrl+Enter Add • Ctrl+L Clear • Ctrl+T Today • Ctrl+O Open • F5 Refresh")
        tip.setStyleSheet("color:#4A6C74;")

        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_clear)
        btn_row.addWidget(self.btn_open)
        btn_row.addWidget(self.btn_export)
        btn_row.addWidget(self.btn_refresh)
        btn_row.addStretch(1)
        btn_row.addWidget(tip)
        root.addLayout(btn_row)

        # Table (3 columns)
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_table_context_menu)
        root.addWidget(self.table)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.update_status(f"File: {EXCEL_PATH}")

        self.setCentralWidget(central)

        # Shortcuts
        self._add_shortcuts()

        # Initial load
        self.load_table()
        self.update_preview()

        # Styling
        self._apply_styles()

    def _apply_styles(self):
        self.setStyleSheet(f"""
            QWidget {{ background: {BG_LIGHT}; color: {TEXT_PRIMARY}; }}
            QRadioButton, QLabel {{ font-size: 12px; }}
            QTextEdit {{ background: white; border: 1px solid #D0E3E6; }}
            QTableWidget {{ background: white; alternate-background-color: {ROW_ALT}; }}
            QPushButton {{ padding: 6px 10px; }}
        """)

    def _add_shortcuts(self):
        def add_seq(seq, handler):
            act = QAction(self)
            act.setShortcut(QKeySequence(seq))
            act.triggered.connect(handler)
            self.addAction(act)

        add_seq("Ctrl+Return", self.on_add)
        add_seq("Ctrl+Enter", self.on_add)
        add_seq("Ctrl+L", self.on_clear)
        add_seq("Ctrl+T", self.set_today)
        add_seq("Ctrl+O", self.open_excel)
        add_seq("F5", self.load_table)

    def set_today(self):
        self.date_edit.setDate(QDate.currentDate())

    def on_clear(self):
        self.desc_text.clear()
        # Do not clear date; we handle Request Number after adding
        self.update_preview()

    def on_add(self):
        sheet = "CR" if self.rb_cr.isChecked() else "WP"

        # Convert QDate to Python date
        dqt = self.date_edit.date()
        d_py = date(dqt.year(), dqt.month(), dqt.day())

        req_num = self.req_edit.text().strip()
        desc = normalize_description(self.desc_text.toPlainText())

        if not desc:
            QMessageBox.critical(self, "Missing description", "Please enter the Description of Work.")
            return

        # If you want to make Request Number mandatory, uncomment:
        # if not req_num:
        #     QMessageBox.critical(self, "Missing request number", "Please enter the Request Number.")
        #     return

        try:
            append_row(EXCEL_PATH, sheet, d_py, req_num, desc)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add entry:\n{e}")
            return

        self.update_status(f"Added to '{sheet}': {d_py.strftime('%Y-%m-%d')}")

        # Prepare for next entry:
        # 1) Clear description
        self.on_clear()
        # 2) Clear Request Number to avoid accidental reuse
        self.req_edit.clear()
        self.req_edit.setFocus()
        # 3) Refresh table
        self.load_table()

    def on_sheet_changed(self):
        self.load_table()

    def update_preview(self):
        text = self.desc_text.toPlainText().strip()
        combined = normalize_description(text)
        self.preview.setText(combined or "(nothing yet)")

    def load_table(self):
        sheet = "CR" if self.rb_cr.isChecked() else "WP"
        try:
            ensure_workbook_and_sheets(EXCEL_PATH)
            rows = read_rows(EXCEL_PATH, sheet)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel:\n{e}")
            return

        self.table.setRowCount(0)
        for dstr, req, desc in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, QTableWidgetItem(dstr))
            self.table.setItem(r, 1, QTableWidgetItem(req))
            self.table.setItem(r, 2, QTableWidgetItem(desc))

        self.update_status(f"File: {EXCEL_PATH} • {sheet} records: {len(rows)}")

    def on_table_context_menu(self, pos: QPoint):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        self.table.selectRow(idx.row())
        menu = QMenu(self)
        act_copy = QAction("Copy row", self)
        act_copy.triggered.connect(self.copy_selected_row)
        menu.addAction(act_copy)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def copy_selected_row(self):
        row = self.table.currentRow()
        if row < 0:
            return
        d = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
        req = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        desc = self.table.item(row, 2).text() if self.table.item(row, 2) else ""
        text = f"{d}\t{req}\t{desc}"
        QApplication.clipboard().setText(text)
        self.update_status("Row copied to clipboard.")

    # -------- Export to OneDrive --------

    def on_export_to_onedrive(self):
        onedrive_dir = find_onedrive_dir()
        default_dir = str(onedrive_dir)
        dest_dir = QFileDialog.getExistingDirectory(self, "Select OneDrive folder to export", default_dir)
        if not dest_dir:
            return
        dest_path = Path(dest_dir) / FILE_NAME

        if dest_path.exists():
            resp = QMessageBox.question(
                self, "Overwrite file?",
                f"'{dest_path}' already exists.\nDo you want to overwrite it?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if resp != QMessageBox.Yes:
                return

        try:
            tmp = dest_path.with_suffix(dest_path.suffix + f".tmp.{os.getpid()}")
            shutil.copy2(EXCEL_PATH, tmp)
            os.replace(tmp, dest_path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", f"Could not export to OneDrive:\n{e}")
            return

        self.update_status(f"Exported to {dest_path}")

    # -------- Common --------

    def open_excel(self):
        path = str(EXCEL_PATH)
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", path], check=False)
            else:
                import subprocess
                subprocess.run(["xdg-open", path], check=False)
        except Exception as e:
            QMessageBox.critical(self, "Open failed", f"Could not open the file:\n{e}")

    def update_status(self, text: str):
        self.status.showMessage(text)


def main():
    # Windows: set AppUserModelID for proper taskbar icon/grouping
    if sys.platform == "win32":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("NetworkChanges.Tracker.1.2")
        except Exception:
            pass

    ensure_workbook_and_sheets(EXCEL_PATH)

    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)

    ico = load_window_icon()
    app.setWindowIcon(ico)

    w = MainWindow()
    w.setWindowIcon(ico)
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
