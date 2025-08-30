import os
import sys
import ctypes
from datetime import date, datetime
from pathlib import Path
from typing import List, Tuple

from PySide6.QtCore import Qt, QDate, QPoint
from PySide6.QtGui import QAction, QIcon, QPainter, QPixmap, QColor, QKeySequence
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QRadioButton, QButtonGroup, QDateEdit, QTextEdit, QTableWidget, QTableWidgetItem,
    QAbstractItemView, QHeaderView, QMessageBox, QMenu, QStatusBar, QFrame
)

# Excel
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("This tool requires 'openpyxl'. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

APP_TITLE = "Network Changes Tracker"
FILE_NAME = "network_changes.xlsx"
SHEETS = ("OCRS", "WP")
HEADERS = ("Approval Date", "Description of Work")
DATE_NUMBER_FORMAT = "yyyy-mm-dd"

# UI colours
ACCENT = "#0E9AA7"
ACCENT_DARK = "#0B7D86"
BG_LIGHT = "#F6FAFB"
ROW_ALT = "#F2F8F9"
TEXT_PRIMARY = "#0B3C49"


def is_frozen() -> bool:
    return getattr(sys, "frozen", False)


def resource_path(*relative) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, *relative)


def app_dir() -> Path:
    if is_frozen():
        return Path(sys.executable).resolve().parent
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


EXCEL_PATH = app_dir() / FILE_NAME


def ensure_workbook_and_sheets(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb["Sheet"])
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws["A1"] = HEADERS[0]
            ws["B1"] = HEADERS[1]
        else:
            ws = wb[sheet_name]
            if (ws["A1"].value is None) and (ws["B1"].value is None):
                ws["A1"] = HEADERS[0]
                ws["B1"] = HEADERS[1]
    wb.save(path)
    return wb


def normalize_description(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = [line.strip() for line in text.split("\n")]
    parts = [p for p in parts if p]
    return ", ".join(parts)


def parse_date_str(s: str) -> date:
    s = (s or "").strip()
    if not s:
        return date.today()
    candidate = s.replace("/", "-")
    parts = candidate.split("-")
    if len(parts) == 3:
        try:
            if len(parts[0]) == 4:  # YYYY-MM-DD
                y, m, d = map(int, parts)
            else:  # DD-MM-YYYY
                d, m, y = map(int, parts)
            return date(y, m, d)
        except Exception as e:
            try:
                return datetime.fromisoformat(s).date()
            except Exception:
                raise ValueError(f"Invalid date format: {s}") from e
    try:
        return datetime.fromisoformat(s).date()
    except Exception as e:
        raise ValueError(f"Invalid date format: {s}") from e


def append_row(path: Path, sheet_name: str, d: date, desc: str):
    wb = ensure_workbook_and_sheets(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws["A1"] = HEADERS[0]
        ws["B1"] = HEADERS[1]
    ws = wb[sheet_name]
    ws.append([d, desc])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).number_format = DATE_NUMBER_FORMAT
    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            f"Cannot save the Excel file.\n\nFile may be open or folder not writable:\n{path}\n\n"
            "Close the file if open, or move the EXE and Excel to a writable folder (e.g., Desktop/Documents)."
        )


def read_rows(path: Path, sheet_name: str) -> List[Tuple[str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows: List[Tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        approval_date, description = row[0], row[1]
        if approval_date is None and description is None:
            continue
        if isinstance(approval_date, (datetime, date)):
            dstr = approval_date.strftime("%Y-%m-%d")
        elif approval_date:
            dstr = str(approval_date)
        else:
            dstr = ""
        rows.append((dstr, description or ""))
    return rows


def emoji_icon(emoji: str, size: int = 128,
               bg=QColor(14, 154, 167), fg=QColor(255, 255, 255)) -> QIcon:
    pm = QPixmap(size, size)
    pm.fill(Qt.transparent)
    painter = QPainter(pm)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setBrush(bg)
    painter.setPen(Qt.NoPen)
    painter.drawEllipse(0, 0, size, size)
    painter.setPen(fg)
    painter.drawText(pm.rect(), Qt.AlignCenter, emoji)
    painter.end()
    return QIcon(pm)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.setMinimumSize(900, 650)
        self.setWindowIcon(emoji_icon("📡"))

        ensure_workbook_and_sheets(EXCEL_PATH)

        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # Banner
        banner = QFrame()
        banner.setStyleSheet(f"background:{ACCENT}; border-radius:6px;")
        bl = QVBoxLayout(banner)
        title = QLabel(APP_TITLE)
        subtitle = QLabel("Track network changes quickly and consistently")
        for w in (title, subtitle):
            w.setStyleSheet("color:white;")
        title.setStyleSheet("color:white; font-size:18px; font-weight:600;")
        subtitle.setStyleSheet("color:#E6FFFF; font-size:12px;")
        bl.addWidget(title)
        bl.addWidget(subtitle)
        root.addWidget(banner)

        # Top row: tracker + date
        top = QHBoxLayout()
        lbl_tracker = QLabel("Tracker:")
        self.rb_ocrs = QRadioButton("OCRS")
        self.rb_wp = QRadioButton("WP")
        self.rb_ocrs.setChecked(True)
        self.sheet_group = QButtonGroup(self)
        self.sheet_group.addButton(self.rb_ocrs)
        self.sheet_group.addButton(self.rb_wp)
        self.sheet_group.buttonToggled.connect(self.on_sheet_changed)

        top.addWidget(lbl_tracker)
        top.addWidget(self.rb_ocrs)
        top.addWidget(self.rb_wp)
        top.addSpacing(20)

        lbl_date = QLabel("Approval Date:")
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        btn_today = QPushButton("Today")
        btn_today.clicked.connect(self.set_today)

        top.addWidget(lbl_date)
        top.addWidget(self.date_edit)
        top.addWidget(btn_today)
        top.addStretch(1)
        root.addLayout(top)

        # Description
        desc_frame = QVBoxLayout()
        lbl_desc = QLabel("Description of Work (multi-line allowed)")
        self.desc_text = QTextEdit()
        self.desc_text.textChanged.connect(self.update_preview)
        self.desc_text.setPlaceholderText("Enter work description; multiple lines will be joined with commas")
        desc_frame.addWidget(lbl_desc)
        desc_frame.addWidget(self.desc_text)
        root.addLayout(desc_frame)

        # Preview
        lbl_prev = QLabel("Preview (single line):")
        self.preview = QLabel("(nothing yet)")
        self.preview.setStyleSheet("background:white; border:1px solid #D0E3E6; padding:8px;")
        root.addWidget(lbl_prev)
        root.addWidget(self.preview)

        # Buttons
        btn_row = QHBoxLayout()
        self.btn_add = QPushButton("Add")
        self.btn_add.clicked.connect(self.on_add)
        self.btn_clear = QPushButton("Clear")
        self.btn_clear.clicked.connect(self.on_clear)
        self.btn_open = QPushButton("Open Excel")
        self.btn_open.clicked.connect(self.open_excel)
        self.btn_refresh = QPushButton("Refresh List")
        self.btn_refresh.clicked.connect(self.load_table)

        tip = QLabel("Shortcuts: Ctrl+Enter Add • Ctrl+L Clear • Ctrl+T Today • Ctrl+O Open • F5 Refresh")
        tip.setStyleSheet("color:#4A6C74;")

        btn_row.addWidget(self.btn_add)
        btn_row.addWidget(self.btn_clear)
        btn_row.addWidget(self.btn_open)
        btn_row.addWidget(self.btn_refresh)
        btn_row.addStretch(1)
        btn_row.addWidget(tip)
        root.addLayout(btn_row)

        # Table
        self.table = QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.on_table_context_menu)
        root.addWidget(self.table)

        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.update_status(f"File: {EXCEL_PATH}")

        self.setCentralWidget(central)

        # Shortcuts
        self._add_shortcuts()

        # Initial load
        self.load_table()
        self.update_preview()

        # Styling
        self._apply_styles()

    def _apply_styles(self):
        self.setStyleSheet(f"""
            QWidget {{ background: {BG_LIGHT}; color: {TEXT_PRIMARY}; }}
            QRadioButton, QLabel {{ font-size: 12px; }}
            QTextEdit {{ background: white; border: 1px solid #D0E3E6; }}
            QTableWidget {{ background: white; alternate-background-color: {ROW_ALT}; }}
            QPushButton {{
                padding: 6px 10px;
            }}
        """)

    def _add_shortcuts(self):
        def add_seq(seq, handler):
            act = QAction(self)
            act.setShortcut(QKeySequence(seq))
            act.triggered.connect(handler)
            self.addAction(act)

        add_seq("Ctrl+Return", self.on_add)
        add_seq("Ctrl+Enter", self.on_add)
        add_seq("Ctrl+L", self.on_clear)
        add_seq("Ctrl+T", self.set_today)
        add_seq("Ctrl+O", self.open_excel)
        add_seq("F5", self.load_table)

    def set_today(self):
        self.date_edit.setDate(QDate.currentDate())

    def on_clear(self):
        self.desc_text.clear()
        self.update_preview()

    def on_add(self):
        sheet = "OCRS" if self.rb_ocrs.isChecked() else "WP"

        # Use QDateEdit value
        d_py = self.date_edit.date().toPython()

        raw = self.desc_text.toPlainText()
        desc = normalize_description(raw)
        if not desc:
            QMessageBox.critical(self, "Missing description", "Please enter the Description of Work.")
            return

        try:
            append_row(EXCEL_PATH, sheet, d_py, desc)
        except PermissionError as e:
            QMessageBox.critical(self, "Cannot save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add entry:\n{e}")
            return

        self.update_status(f"Added to '{sheet}': {d_py.strftime('%Y-%m-%d')}")
        self.on_clear()
        self.load_table()

    def on_sheet_changed(self):
        self.load_table()

    def update_preview(self):
        text = self.desc_text.toPlainText().strip()
        combined = normalize_description(text)
        self.preview.setText(combined or "(nothing yet)")

    def load_table(self):
        sheet = "OCRS" if self.rb_ocrs.isChecked() else "WP"
        try:
            ensure_workbook_and_sheets(EXCEL_PATH)
            rows = read_rows(EXCEL_PATH, sheet)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read Excel:\n{e}")
            return

        self.table.setRowCount(0)
        for dstr, desc in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            item_date = QTableWidgetItem(dstr)
            item_desc = QTableWidgetItem(desc)
            self.table.setItem(r, 0, item_date)
            self.table.setItem(r, 1, item_desc)

        self.update_status(f"File: {EXCEL_PATH} • {sheet} records: {len(rows)}")

    def on_table_context_menu(self, pos: QPoint):
        idx = self.table.indexAt(pos)
        if not idx.isValid():
            return
        self.table.selectRow(idx.row())
        menu = QMenu(self)
        act_copy = QAction("Copy row", self)
        act_copy.triggered.connect(self.copy_selected_row)
        menu.addAction(act_copy)
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def copy_selected_row(self):
        row = self.table.currentRow()
        if row < 0:
            return
        d = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
        desc = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        text = f"{d}\t{desc}"
        QApplication.clipboard().setText(text)
        self.update_status("Row copied to clipboard.")

    def open_excel(self):
        path = str(EXCEL_PATH)
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", path], check=False)
            else:
                import subprocess
                subprocess.run(["xdg-open", path], check=False)
        except Exception as e:
            QMessageBox.critical(self, "Open failed", f"Could not open the file:\n{e}")

    def update_status(self, text: str):
        self.status.showMessage(text)


def main():
    # Windows: set AppUserModelID for proper taskbar icon/grouping
    if sys.platform == "win32":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("NetworkChanges.Tracker.1.0")
        except Exception:
            pass

    ensure_workbook_and_sheets(EXCEL_PATH)

    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setWindowIcon(emoji_icon("📡"))

    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
